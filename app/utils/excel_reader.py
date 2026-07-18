from openpyxl import load_workbook


class ExcelReader:
    @staticmethod
    def read_data(file_path, sheet_name=None):
        """
        Read test data from an Excel file.

        Args:
            file_path (str): Path to the Excel file.
            sheet_name (str, optional): Sheet name. Uses active sheet if None.

        Returns:
            list[dict]: List of dictionaries containing row data.
        """

        workbook = load_workbook(file_path)

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))

        headers = rows[0]
        data = []

        for row in rows[1:]:
            data.append(dict(zip(headers, row)))

        workbook.close()

        return data